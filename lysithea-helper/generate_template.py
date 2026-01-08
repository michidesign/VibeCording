#!/usr/bin/env python3
"""Excelテンプレート生成スクリプト"""

import subprocess
import sys

# openpyxlをインストール（なければ）
try:
    import openpyxl
except ImportError:
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ワークブック作成
wb = Workbook()

# マスターシート（JOBコード一覧）
ws_master = wb.active
ws_master.title = 'マスター'

job_codes = [
    '1364042000:ＤＩ・ＤＣ１統・ＤＭＣ３・２課',
    '1364042016:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・未稼働費用',
    '1364042017:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・社内業務',
    '1364042018:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・社内会議',
    '1364042019:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・ＭＧ費用',
    '1364042023:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・キャリアモデルバッジ制度費用',
    '1364042024:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・その他研修',
    '1364042025:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・既存プリセールス',
    '1364042026:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・新規プリセールス',
    '2103405001:（名）ＴＭ／オンサイト（ＳＰ部ＷＥＢ制作開発室）',
    '2112671029:レッドバロン／サイト運用',
    '2113906001:（名）オークローンマーケティング／サイト運用',
    '2127301001:（名オ）名古屋テレビ／サイト運用',
    '2143929036:ＮＴＴドコモ／Ｌｅｍｉｎｏ案件',
    '2162938001:【３Ｇ】ＮＴＴドコモｄカード／年間プロモーション業務',
    '2166827002:ＵＣＢ／ＳＦＭＣ運用（２４年４月～）',
    '2172796001:コープデリ／ＬＩＮＥチャットリタゲ運用（２５年１０月～）',
    '2172997001:テック情報／サイトリニューアル',
    '2173244001:ヴィアトリス製薬／企画コンテンツ制作',
    '2173589001:アイキャスト／ひかりＴＶ運用',
]

ws_master['A1'] = 'JOBコード一覧'
ws_master['A1'].font = Font(bold=True)
for i, code in enumerate(job_codes, start=2):
    ws_master[f'A{i}'] = code
ws_master.column_dimensions['A'].width = 60

# 入力シート
ws_input = wb.create_sheet('勤怠入力')

# スタイル定義
header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
optional_fill = PatternFill(start_color='90CAF9', end_color='90CAF9', fill_type='solid')  # 任意項目用
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ヘッダー（日付列を追加、作業時間は任意）
headers = ['日付', 'JOBコード', '作業開始', '作業終了', '作業時間（任意）', 'コメント']
col_widths = [14, 60, 12, 12, 14, 40]

for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws_input.cell(row=1, column=col, value=header)
    # 任意項目は薄い色で表示
    if '任意' in header:
        cell.fill = optional_fill
    else:
        cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center')
    ws_input.column_dimensions[get_column_letter(col)].width = width

# データ入力行（50行）
for row in range(2, 52):
    for col in range(1, 7):
        cell = ws_input.cell(row=row, column=col)
        cell.border = border
        # 日付列のフォーマット
        if col == 1:
            cell.number_format = 'YYYY/MM/DD'

# プルダウン設定（JOBコード列 = B列）
dv = DataValidation(
    type='list',
    formula1='マスター!$A$2:$A$21',
    allow_blank=True
)
dv.error = 'マスターシートから選択してください'
dv.errorTitle = '入力エラー'
ws_input.add_data_validation(dv)
dv.add('B2:B51')

# サンプルデータ
from datetime import date
ws_input['A2'] = date(2026, 1, 8)
ws_input['B2'] = '2162938001:【３Ｇ】ＮＴＴドコモｄカード／年間プロモーション業務'
ws_input['C2'] = '0900'
ws_input['D2'] = '1200'
ws_input['E2'] = ''  # 作業時間は任意なので空欄
ws_input['F2'] = 'dカード定例MTG'

ws_input['A3'] = date(2026, 1, 8)
ws_input['B3'] = '1364042017:ＤＩ・ＤＣ１統・ＤＭＣ３・２課・社内業務'
ws_input['C3'] = '1300'
ws_input['D3'] = '1750'
ws_input['E3'] = ''  # 作業時間は任意なので空欄
ws_input['F3'] = '資料作成'

# 入力シートをアクティブに
wb.active = ws_input

# 保存
output_path = '/Users/uchiyamatakahiro/Library/CloudStorage/GoogleDrive-danielerlandsson0809@gmail.com/マイドライブ/VibeCording/lysithea-helper/勤怠入力テンプレート.xlsx'
wb.save(output_path)
print(f'✅ Excelテンプレートを作成しました: {output_path}')

